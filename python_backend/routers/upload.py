import os
import uuid
import tempfile
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from typing import Optional

from database import get_db
from models.chat import Message, Attachment
from services.rag_pipeline import ingest_documents_to_pack
from langchain_core.documents import Document

import PyPDF2
import openpyxl
import base64
from groq import Groq

# Reuse the Groq API key from environment
GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
groq_client = Groq(api_key=GROQ_API_KEY) if GROQ_API_KEY else None

router = APIRouter(prefix="/api/upload", tags=["upload"])

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

def parse_pdf(file_path: str) -> str:
    text = ""
    try:
        with open(file_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                extracted = page.extract_text()
                if extracted:
                    text += extracted + "\n"
    except Exception as e:
        print(f"PDF Parse error: {e}")
    return text

def parse_excel(file_path: str) -> str:
    text = ""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text += f"--- Sheet: {sheet_name} ---\n"
            for row in sheet.iter_rows(values_only=True):
                row_str = " | ".join([str(c) if c is not None else "" for c in row])
                if row_str.strip() and row_str.strip() != "|":
                    text += row_str + "\n"
    except Exception as e:
        print(f"Excel Parse error: {e}")
    return text

def parse_text(file_path: str) -> str:
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    except Exception as e:
        print(f"Text Parse error: {e}")
        return ""

def parse_image(file_path: str) -> str:
    if not groq_client:
        return "[Vision AI Disabled: Missing Groq API Key]"
    try:
        with open(file_path, "rb") as f:
            encoded_string = base64.b64encode(f.read()).decode('utf-8')
        
        file_extension = os.path.splitext(file_path)[1].lower().replace('.', '')
        mime_type = f"image/{file_extension}"
        
        completion = groq_client.chat.completions.create(
            model="llama-3.2-11b-vision-preview",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": "Extract all text from this image accurately (OCR). Also, provide a brief description of any UI elements, charts, or errors visible."
                        },
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime_type};base64,{encoded_string}"
                            }
                        }
                    ]
                }
            ],
            temperature=0.1,
            max_tokens=1024,
        )
        return completion.choices[0].message.content or ""
    except Exception as e:
        print(f"Image Vision OCR error: {e}")
        return f"[Image OCR Failed: {e}]"

@router.post("/")
async def upload_file(
    file: UploadFile = File(...),
    message_id: str = Form(...),
    db: Session = Depends(get_db)
):
    """
    Handles a file upload, parses its content, stores it as an Attachment linked to a Message,
    and vectorizes the content into ChromaDB for File Intelligence.
    """
    message = db.query(Message).filter(Message.id == message_id).first()
    if not message:
        raise HTTPException(status_code=404, detail="Message not found")

    file_extension = os.path.splitext(file.filename)[1].lower()
    
    # Save file physically
    file_id = str(uuid.uuid4())
    safe_filename = f"{file_id}{file_extension}"
    file_path = os.path.join(UPLOAD_DIR, safe_filename)
    
    with open(file_path, "wb") as f:
        f.write(await file.read())

    # Extract text based on file type
    extracted_text = ""
    if file_extension == ".pdf":
        extracted_text = parse_pdf(file_path)
    elif file_extension in [".xlsx", ".csv"]:
        extracted_text = parse_excel(file_path)
    elif file_extension in [".txt", ".md", ".py", ".js", ".ts", ".html", ".css", ".java", ".cpp", ".cs", ".php"]:
        extracted_text = parse_text(file_path)
    elif file_extension in [".png", ".jpg", ".jpeg", ".webp"]:
        extracted_text = parse_image(file_path)
    else:
        extracted_text = "[Unsupported file format for text extraction]"

    # Create Attachment record
    attachment = Attachment(
        id=file_id,
        message_id=message_id,
        file_name=file.filename,
        file_type=file_extension,
        file_path=file_path,
        extracted_text=extracted_text
    )
    db.add(attachment)
    db.commit()

    # Vectorize the text for File Intelligence (using Conversation ID as the Pack Name)
    if extracted_text and not extracted_text.startswith("["):
        doc = Document(
            page_content=extracted_text, 
            metadata={"source": file.filename, "file_id": file_id, "conversation_id": message.conversation_id}
        )
        ingest_documents_to_pack([doc], pack_name=f"conv_{message.conversation_id}")

    return {
        "status": "success",
        "attachment_id": file_id,
        "file_name": file.filename,
        "parsed_length": len(extracted_text)
    }
